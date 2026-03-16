"""
src/parser/excel_writer.py
---------------------------
Write parsed DataFrames to individual and master Excel files.

Individual output: ``data/output/parsed/{company}_{year}.xlsx``
Master output:     ``data/output/master_shareholder_data.xlsx``

The master file is deduplicated on (folio_no, source_file) before save
so re-running the parser does not produce duplicate rows.
"""

from __future__ import annotations

import logging
import os
import time
from pathlib import Path

import pandas as pd

_logger = logging.getLogger("parser.excel_writer")


# ---------------------------------------------------------------------------
# Safe Excel Write
# ---------------------------------------------------------------------------

def safe_excel_write(df: pd.DataFrame, filepath: str, **kwargs) -> str:
    """Write DataFrame to Excel with retry logic for locked files.
    
    Attempts to write to a temporary file, then rename to final path.
    If file is locked, retries up to 5 times with 3-second delays.
    If all attempts fail, saves with a timestamp suffix instead of crashing.
    
    Args:
        df: DataFrame to write.
        filepath: Target Excel file path.
        **kwargs: Additional arguments to pass to to_excel().
    
    Returns:
        Actual filepath where data was written.
    """
    max_attempts = 5
    filepath = str(filepath)
    
    for attempt in range(max_attempts):
        try:
            # Write to temp file first (use .xlsx extension to avoid engine issues)
            temp_path = filepath.replace('.xlsx', '_temp.xlsx')
            df.to_excel(temp_path, **kwargs)
            # Then rename to final
            if os.path.exists(filepath):
                os.remove(filepath)
            os.rename(temp_path, filepath)
            _logger.debug(f"Wrote Excel file: {filepath}")
            return filepath
        except (PermissionError, OSError) as e:
            if attempt < max_attempts - 1:
                _logger.debug(f"Excel write attempt {attempt + 1} failed, retrying in 3s: {e}")
                time.sleep(3)
                continue
            else:
                # Save with timestamp instead
                ts = time.strftime('%H%M%S')
                alt_path = filepath.replace('.xlsx', f'_{ts}.xlsx')
                df.to_excel(alt_path, **kwargs)
                _logger.warning(f"File locked after {max_attempts} attempts — saved to {alt_path} instead")
                return alt_path


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

_HEADER_FILL_COLOR = "1A3C5E"   # deep navy (matches Tech Mahindra brand)
_HEADER_FONT_COLOR = "FFFFFF"


def format_excel_output(file_path: str, df: pd.DataFrame) -> None:
    """Apply column formatting and auto-width to an Excel file.
    
    Sets minimum width of 30 for key columns (company_name, source_file, address, name)
    and auto-width up to 60 for others. Also applies header formatting.
    
    Args:
        file_path: Path to the Excel file to format.
        df: DataFrame that was written to the file (for column info).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Header row formatting
        header_fill = PatternFill(
            start_color=_HEADER_FILL_COLOR,
            end_color=_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(color=_HEADER_FONT_COLOR, bold=True, size=10)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-width (cap at 60 chars), with minimum of 30 for key columns
        key_cols_min_width = {'company_name', 'source_file', 'address', 'name'}
        for col_idx, col in enumerate(df.columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0,
            )
            # Set minimum width for key columns, otherwise cap at 60
            min_width = 30 if col in key_cols_min_width else 0
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, 60))
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        wb.save(file_path)
    except Exception as exc:
        _logger.debug("Could not apply Excel formatting: %s", exc)


def _apply_formatting(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Apply column auto-width and header formatting via openpyxl."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415

        ws = writer.sheets[sheet_name]

        # Header row formatting
        header_fill = PatternFill(
            start_color=_HEADER_FILL_COLOR,
            end_color=_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(color=_HEADER_FONT_COLOR, bold=True, size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-width (cap at 60 chars), with minimum of 30 for key columns
        key_cols_min_width = {'company_name', 'source_file', 'address', 'name'}
        for col_idx, col in enumerate(df.columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0,
            )
            # Set minimum width for key columns, otherwise cap at 60
            min_width = 30 if col in key_cols_min_width else 0
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, 60))

        # Freeze top row
        ws.freeze_panes = "A2"

    except Exception as exc:  # noqa: BLE001
        _logger.debug("Could not apply Excel formatting: %s", exc)


# ---------------------------------------------------------------------------
# Public writers
# ---------------------------------------------------------------------------

def write_individual(
    df: pd.DataFrame,
    output_dir: str,
    company: str,
    year: str,
) -> str:
    """Write *df* to a per-company Excel file.

    Args:
        df:         Normalized DataFrame to write.
        output_dir: Directory for per-file output.
        company:    Company slug used in the filename.
        year:       Financial year used in the filename.

    Returns:
        Absolute path to the written ``.xlsx`` file.

    Raises:
        ValueError: If *df* is empty.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Sanitise filename components
    safe_company = company.replace("/", "-").replace("\\", "-")
    safe_year    = year.replace("/", "-")
    out_path = out_dir / f"{safe_company}_{safe_year}.xlsx"

    sheet = "Shareholders"
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        _apply_formatting(writer, sheet, df)

    _logger.info(
        "Wrote individual Excel: %s (%d records)", out_path, len(df)
    )
    return str(out_path)


def append_to_master(
    df: pd.DataFrame,
    master_path: str,
    dedup_cols: list[str] | None = None,
) -> None:
    """Append *df* to the master Excel file, deduplicating as needed.

    If the master file does not yet exist, it is created.
    Deduplication key defaults to ``["folio_no", "source_file"]``.

    Args:
        df:           Normalized DataFrame to append.
        master_path:  Path to the master ``.xlsx`` file.
        dedup_cols:   Columns to use as deduplication key.
    """
    if df is None or df.empty:
        _logger.debug("append_to_master: empty DataFrame, skipping.")
        return

    dedup_cols = dedup_cols or ["folio_no", "source_file"]
    master = Path(master_path)
    master.parent.mkdir(parents=True, exist_ok=True)

    if master.exists():
        try:
            existing = pd.read_excel(master, engine="openpyxl", dtype=str)
            combined = pd.concat([existing, df.astype(str)], ignore_index=True)
        except PermissionError:
            # File may be locked by another process - add delay and retry once
            import time
            _logger.warning("Master file locked, retrying in 2 seconds...")
            time.sleep(2)
            try:
                existing = pd.read_excel(master, engine="openpyxl", dtype=str)
                combined = pd.concat([existing, df.astype(str)], ignore_index=True)
            except PermissionError as exc2:
                _logger.error("Master file still locked after retry: %s", exc2)
                return
        except Exception as exc:  # noqa: BLE001
            _logger.warning(
                "Could not read existing master %s (%s). Overwriting.", master, exc
            )
            combined = df.copy()
    else:
        combined = df.copy()

    # Deduplicate
    valid_dedup = [c for c in dedup_cols if c in combined.columns]
    if valid_dedup:
        before = len(combined)
        combined = combined.drop_duplicates(subset=valid_dedup, keep="last")
        dropped = before - len(combined)
        if dropped:
            _logger.debug("Master dedup: removed %d duplicate rows.", dropped)

    combined = combined.sort_values(
        ["company_name", "year", "folio_no"],
        na_position="last",
    ).reset_index(drop=True)

    sheet = "Master"
    try:
        with pd.ExcelWriter(str(master), engine="openpyxl") as writer:
            combined.to_excel(writer, index=False, sheet_name=sheet)
            _apply_formatting(writer, sheet, combined)

        _logger.info(
            "Master Excel updated: %s (%d total records)", master, len(combined)
        )
    except PermissionError as exc:
        _logger.error("Failed to append to master Excel: %s", exc)
    except Exception as exc:  # noqa: BLE001
        _logger.error("Error writing master Excel: %s", exc)


# ---------------------------------------------------------------------------
# REQUIREMENT 3: Multi-sheet Excel generation
# ---------------------------------------------------------------------------

def create_multisheet_excel(df: pd.DataFrame, output_path: str) -> str:
    """
    REQUIREMENT 3: Create master Excel with ALL_COMPANIES sheet + company-specific sheets.
    
    REQUIREMENT 1: Replace all NaN with empty string (for Excel output).
    
    Args:
        df: Merged DataFrame with all records
        output_path: Path to output Excel file
        
    Returns:
        Path to created file
    """
    if df.empty:
        _logger.warning("No data to write to Excel")
        return output_path
    
    # Final NaN cleanup: convert all columns to avoid blank-cell round-trip issues.
    # Numeric cols → 0 for NaN; all other cols → "" for NaN, cast to object so
    # openpyxl writes actual empty strings instead of blank cells.
    df_clean = df.copy()
    for col in df_clean.columns:
        dtype_str = str(df_clean[col].dtype)
        if 'float' in dtype_str or 'int' in dtype_str:
            # Check if column is all-null (would be all-NaN float64 from read_excel)
            if df_clean[col].isna().all():
                df_clean[col] = df_clean[col].fillna("").astype(object)
            else:
                df_clean[col] = df_clean[col].fillna(0)
        else:
            df_clean[col] = df_clean[col].fillna("").astype(object).replace({"nan": "", "<NA>": ""})
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: ALL_COMPANIES - all records with all columns
            safe_sheet_name = _sanitize_sheet_name("ALL_COMPANIES")
            df_clean.to_excel(writer, sheet_name=safe_sheet_name, index=False, na_rep="")
            _logger.info(f"Created sheet: {safe_sheet_name} ({len(df_clean)} records)")
            
            # Company-specific sheets
            companies = df_clean['company_name'].unique() if 'company_name' in df_clean.columns else []
            
            for company in companies:
                if pd.isna(company) or str(company).strip() == "":
                    continue
                
                company_df = df_clean[df_clean['company_name'] == str(company)].copy()
                sheet_name = _sanitize_sheet_name(str(company).upper())
                
                try:
                    company_df.to_excel(writer, sheet_name=sheet_name, index=False, na_rep="")
                    _logger.info(f"Created sheet: {sheet_name} ({len(company_df)} records)")
                except Exception as e:
                    _logger.error(f"Failed to create sheet for {company}: {e}")
        
        _logger.info(f"Multi-sheet Excel created: {output_path}")
        return output_path
        
    except Exception as e:
        _logger.error(f"Failed to create multi-sheet Excel: {e}")
        raise


def _sanitize_sheet_name(name: str) -> str:
    """
    Sanitize sheet name for Excel (max 31 chars, no special chars).
    
    Args:
        name: Raw sheet name
        
    Returns:
        Sanitized sheet name
    """
    # Remove special characters
    sanitized = "".join(c if c.isalnum() or c in "_-" else "_" for c in str(name))
    # Limit to 31 characters (Excel limit)
    sanitized = sanitized[:31]
    return sanitized or "Sheet"
